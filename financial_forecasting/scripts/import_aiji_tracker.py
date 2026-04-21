"""AIJI Project Tracker v12 -> bedrock.project tree importer.

Loads the Excel tracker as the new source of truth for the AIJI project.
Soft-deletes existing workstream/milestone/task rows (reversible via
POST /api/projects/{id}/restore) then inserts a clean tree in one transaction.
Mirrors the merge/cascade semantics of
routes/projects.py::import_project_data so the import honours the same
FK, CHECK, and updated_at-trigger rules.

Usage (from the financial_forecasting directory):

    python -m scripts.import_aiji_tracker \\
        --input "/Users/jp/Downloads/AIJI Project Tracker_v12.xlsx" \\
        --project-id a0000000-0000-4000-8000-000000000001 \\
        --imported-by jp.dery@pursuit.org \\
        [--dry-run] [--yes] [--update-project-desc]

DATABASE_URL is required for live runs (same discipline as db.py).
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import os
import re
import sys
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse

import asyncpg
import openpyxl
from dotenv import load_dotenv

from scripts._owner_parser import build_user_map, parse_owner_string

# Match main.py's env loading so DATABASE_URL from financial_forecasting/.env
# is picked up without the caller having to export it manually.
load_dotenv(override=False)

logger = logging.getLogger("import_aiji_tracker")

# ---- Schema constants (mirror financial_forecasting/db/init.sql) ----

MILESTONE_STATUS_ENUM = {"On Track", "At Risk", "Needs Attention", "Completed"}
MILESTONE_PRIORITY_ENUM = {"Now", "Later", "On-going"}
TASK_STATUS_ENUM = {"Not Started", "In Progress", "Completed", "Blocked", "On Hold"}


def _norm(value: Any) -> str:
    """Lowercase, trim, collapse internal whitespace — for dict-key matching."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


# ---- Excel -> DB mappings (rationale in the approved plan) ----

MILESTONE_STATUS_MAP: Dict[str, str] = {
    _norm("On Track"): "On Track",
    _norm("On-Track"): "On Track",
    _norm("At Risk"): "At Risk",
    _norm("Needs Attention"): "Needs Attention",
    _norm("Completed"): "Completed",
    _norm("Complete"): "Completed",
    _norm("Done"): "Completed",
    # Milestone enum has no "Not Started"; callers pair with priority="Later"
    # to retain the upcoming-not-started signal without violating the CHECK.
    _norm("Not Started"): "On Track",
}

TASK_STATUS_MAP: Dict[str, str] = {
    _norm("Not Started"): "Not Started",
    _norm("In Progress"): "In Progress",
    # Task enum has no "On Track"; Excel uses it as "actively worked".
    _norm("On Track"): "In Progress",
    _norm("Completed"): "Completed",
    _norm("Complete"): "Completed",
    _norm("Done"): "Completed",
    _norm("Blocked"): "Blocked",
    _norm("On Hold"): "On Hold",
    _norm("Paused"): "On Hold",
}

TIMELINE_PRIORITY_MAP: Dict[str, str] = {
    _norm("April - May"): "Now",
    _norm("April-May"): "Now",
    _norm("April May"): "Now",
    _norm("May - June"): "Now",
    _norm("June - July"): "Later",
    _norm("June-July"): "Later",
    _norm("July - August"): "Later",
    _norm("Aug - Sept"): "Later",
    _norm("Aug-Sept"): "Later",
    _norm("August - September"): "Later",
    _norm("Ongoing"): "On-going",
    _norm("On-going"): "On-going",
    _norm("On going"): "On-going",
}

WORKSTREAM_TYPO_FIX: Dict[str, str] = {
    _norm("3. Parnterships & Development"): "3. Partnerships & Development",
}

URL_RE = re.compile(r"https?://[^\s,;<>\"']+", re.IGNORECASE)

DEFAULT_PROJECT_ID = "a0000000-0000-4000-8000-000000000001"
DEFAULT_IMPORTED_BY = "jp.dery@pursuit.org"
DEFAULT_INPUT_PATH = "/Users/jp/Downloads/AIJI Project Tracker_v12.xlsx"

SHEET_ABOUT = "About "
SHEET_MILESTONES = "Milestones"
SHEET_TASKS = "Tasks"

UPDATED_PROJECT_DESCRIPTION = (
    "AI Jobs Institute — Pursuit's flagship workforce initiative. Tracker covers "
    "Strategy & Design, Construction & Financing, Partnerships & Development, "
    "Program Delivery, and Launch & Activation."
)


# ---- In-memory tree ----


@dataclass
class TaskRow:
    title: str
    status: str
    owner: str
    deadline: Optional[date]
    start_date: Optional[date]
    description: str
    updates: str
    links: List[str]
    sort_order: int
    owner_ids: List[uuid.UUID] = field(default_factory=list)


@dataclass
class MilestoneRow:
    title: str
    status: str
    priority: str
    owner: str
    description: str
    source_links: List[str]
    sort_order: int
    tasks: List[TaskRow] = field(default_factory=list)
    owner_ids: List[uuid.UUID] = field(default_factory=list)


@dataclass
class WorkstreamRow:
    name: str
    description: str
    sort_order: int
    milestones: List[MilestoneRow] = field(default_factory=list)


@dataclass
class ParsedTree:
    workstreams: List[WorkstreamRow]
    warnings: List[str]


# ---- Cell + URL helpers ----


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_workstream_name(raw: Any) -> str:
    clean = _cell(raw)
    if not clean:
        return ""
    return WORKSTREAM_TYPO_FIX.get(_norm(clean), clean)


def extract_urls(raw: Any) -> List[str]:
    if raw is None:
        return []
    urls = URL_RE.findall(str(raw))
    return [u.rstrip(".,;:)\"'") for u in urls]


def extract_non_url_text(raw: Any) -> str:
    if raw is None:
        return ""
    stripped = URL_RE.sub(" ", str(raw))
    stripped = re.sub(r"\s+", " ", stripped).strip()
    return re.sub(r"^[\s,;:]+|[\s,;:]+$", "", stripped)


def to_date(val: Any) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        try:
            return datetime.fromisoformat(s).date()
        except ValueError:
            return None
    return None


# ---- Status / priority mapping ----


def map_milestone_status(raw: Any, warnings: List[str]) -> str:
    key = _norm(raw)
    if not key:
        return "On Track"
    if key in MILESTONE_STATUS_MAP:
        return MILESTONE_STATUS_MAP[key]
    warnings.append(f"Unknown milestone status {raw!r}; defaulting to 'On Track'")
    return "On Track"


def map_task_status(raw: Any, warnings: List[str]) -> str:
    key = _norm(raw)
    if not key:
        return "Not Started"
    if key in TASK_STATUS_MAP:
        return TASK_STATUS_MAP[key]
    warnings.append(f"Unknown task status {raw!r}; defaulting to 'Not Started'")
    return "Not Started"


def map_timeline_to_priority(raw: Any, is_not_started: bool, warnings: List[str]) -> str:
    key = _norm(raw)
    if not key:
        return "Later" if is_not_started else "Now"
    mapped = TIMELINE_PRIORITY_MAP.get(key)
    if mapped:
        return mapped
    if is_not_started:
        return "Later"
    warnings.append(f"Unknown timeline {raw!r}; defaulting to 'Now'")
    return "Now"


_WS_PREFIX_RE = re.compile(r"^\s*(\d+)[.)]\s*")


def _workstream_sort_order(name: str, fallback: int) -> int:
    m = _WS_PREFIX_RE.match(name or "")
    if m:
        try:
            return int(m.group(1)) - 1
        except ValueError:
            return fallback
    return fallback


# ---- Excel parsing ----


def parse_about(wb: openpyxl.Workbook) -> Dict[str, str]:
    """Return {normalized_workstream_name: description}."""
    if SHEET_ABOUT not in wb.sheetnames:
        return {}
    ws = wb[SHEET_ABOUT]
    out: Dict[str, str] = {}
    in_block = False
    for row in ws.iter_rows(values_only=True):
        if len(row) < 3:
            continue
        label = _cell(row[1])
        value = _cell(row[2])
        if label.lower() == "workstream" and value.lower() == "purpose":
            in_block = True
            continue
        if not in_block:
            continue
        if label.lower() == "scope" or label.lower() == "what this tracker is":
            break
        if not label:
            continue
        if value:
            clean = normalize_workstream_name(label)
            out[_norm(clean)] = value
    return out


def parse_milestones(
    wb: openpyxl.Workbook,
    about: Dict[str, str],
    warnings: List[str],
) -> Dict[str, WorkstreamRow]:
    ws = wb[SHEET_MILESTONES]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, r in enumerate(rows):
        if len(r) >= 1 and _cell(r[0]).lower() == "workstream":
            header_idx = i
            break
    if header_idx is None:
        raise RuntimeError(f"Could not find header row in sheet {SHEET_MILESTONES!r}")

    ws_map: Dict[str, WorkstreamRow] = {}
    ms_sort_by_ws: Dict[str, int] = {}

    for r in rows[header_idx + 1 :]:
        if not any(_cell(c) for c in r):
            continue
        ws_name = normalize_workstream_name(r[0] if len(r) > 0 else None)
        ms_title = _cell(r[1] if len(r) > 1 else None)
        if not ws_name or not ms_title:
            continue
        status_raw = r[2] if len(r) > 2 else None
        timeline_raw = r[3] if len(r) > 3 else None
        desc_raw = r[4] if len(r) > 4 else None
        owner_raw = r[5] if len(r) > 5 else None
        ext_raw = r[6] if len(r) > 6 else None
        source_raw = r[7] if len(r) > 7 else None

        is_not_started = _norm(status_raw) == _norm("Not Started")
        status = map_milestone_status(status_raw, warnings)
        priority = map_timeline_to_priority(timeline_raw, is_not_started, warnings)

        parts: List[str] = []
        desc_text = _cell(desc_raw)
        if desc_text:
            parts.append(desc_text)
        ext_text = _cell(ext_raw)
        if ext_text:
            parts.append(f"External support: {ext_text}")
        non_url_source = extract_non_url_text(source_raw)
        if non_url_source:
            parts.append(f"Source: {non_url_source}")
        description = "\n\n".join(parts)
        source_links = extract_urls(source_raw)

        ws_key = _norm(ws_name)
        if ws_key not in ws_map:
            ws_map[ws_key] = WorkstreamRow(
                name=ws_name,
                description=about.get(ws_key, ""),
                sort_order=_workstream_sort_order(ws_name, len(ws_map)),
            )
            ms_sort_by_ws[ws_key] = 0

        sort_order = ms_sort_by_ws[ws_key]
        ms_sort_by_ws[ws_key] = sort_order + 1
        ws_map[ws_key].milestones.append(
            MilestoneRow(
                title=ms_title,
                status=status,
                priority=priority,
                owner=_cell(owner_raw),
                description=description,
                source_links=source_links,
                sort_order=sort_order,
            )
        )

    return ws_map


def parse_tasks(
    wb: openpyxl.Workbook,
    ws_map: Dict[str, WorkstreamRow],
    warnings: List[str],
) -> None:
    if SHEET_TASKS not in wb.sheetnames:
        warnings.append(f"Workbook has no {SHEET_TASKS!r} sheet; skipping tasks")
        return
    ws = wb[SHEET_TASKS]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, r in enumerate(rows):
        vals = [_cell(c).lower() for c in r]
        if "task" in vals[:3] and "status" in vals[:5]:
            header_idx = i
            break
    if header_idx is None:
        raise RuntimeError(f"Could not find header row in sheet {SHEET_TASKS!r}")

    ms_by_key: Dict[str, Dict[str, MilestoneRow]] = {
        ws_key: {_norm(ms.title): ms for ms in ws_row.milestones}
        for ws_key, ws_row in ws_map.items()
    }

    current_ws_key: Optional[str] = None
    current_ms: Optional[MilestoneRow] = None
    task_sort_by_ms: Dict[str, int] = {}

    for r in rows[header_idx + 1 :]:
        col_a = _cell(r[0]) if len(r) > 0 else ""
        col_b = _cell(r[1]) if len(r) > 1 else ""
        col_c = _cell(r[2]) if len(r) > 2 else ""

        if col_a and not col_b and not col_c:
            ws_name = normalize_workstream_name(col_a)
            current_ws_key = _norm(ws_name)
            current_ms = None
            if current_ws_key not in ws_map:
                warnings.append(
                    f"Tasks sheet references unknown workstream {ws_name!r}; "
                    "tasks under it will be skipped"
                )
                current_ws_key = None
            continue

        if col_b and not col_c:
            current_ms = None
            if current_ws_key is None:
                warnings.append(
                    f"Milestone header {col_b!r} seen before a valid workstream; skipping"
                )
                continue
            ms = ms_by_key.get(current_ws_key, {}).get(_norm(col_b))
            if ms is None:
                warnings.append(
                    f"Tasks sheet references unknown milestone {col_b!r} "
                    f"in workstream {ws_map[current_ws_key].name!r}; tasks under it skipped"
                )
            else:
                current_ms = ms
            continue

        if not col_c:
            continue

        if current_ws_key is None or current_ms is None:
            warnings.append(f"Task {col_c!r} seen without valid workstream/milestone; skipping")
            continue

        status_raw = r[3] if len(r) > 3 else None
        owner_raw = r[4] if len(r) > 4 else None
        deadline_raw = r[5] if len(r) > 5 else None
        desc_raw = r[6] if len(r) > 6 else None
        updates_raw = r[7] if len(r) > 7 else None
        link_raw = r[8] if len(r) > 8 else None

        status = map_task_status(status_raw, warnings)
        owner = _cell(owner_raw)
        deadline = to_date(deadline_raw)
        start_date = deadline - timedelta(days=7) if deadline else None

        base_desc = _cell(desc_raw)
        non_url_link = extract_non_url_text(link_raw)
        desc_parts = [p for p in (base_desc, f"Link: {non_url_link}" if non_url_link else "") if p]
        description = "\n\n".join(desc_parts)
        links = extract_urls(link_raw)
        updates = _cell(updates_raw)

        key = f"{current_ws_key}::{_norm(current_ms.title)}"
        sort_order = task_sort_by_ms.get(key, 0)
        task_sort_by_ms[key] = sort_order + 1

        current_ms.tasks.append(
            TaskRow(
                title=col_c,
                status=status,
                owner=owner,
                deadline=deadline,
                start_date=start_date,
                description=description,
                updates=updates,
                links=links,
                sort_order=sort_order,
            )
        )


def parse_workbook(path: Path) -> ParsedTree:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    warnings: List[str] = []
    about = parse_about(wb)
    ws_map = parse_milestones(wb, about, warnings)
    parse_tasks(wb, ws_map, warnings)
    workstreams = sorted(ws_map.values(), key=lambda w: (w.sort_order, w.name))

    for ws_row in workstreams:
        if ws_row.sort_order < 0:
            raise RuntimeError(f"Negative sort_order for workstream {ws_row.name!r}")
        for ms in ws_row.milestones:
            if ms.status not in MILESTONE_STATUS_ENUM:
                raise RuntimeError(
                    f"Milestone status {ms.status!r} violates CHECK (title={ms.title!r})"
                )
            if ms.priority not in MILESTONE_PRIORITY_ENUM:
                raise RuntimeError(
                    f"Milestone priority {ms.priority!r} violates CHECK (title={ms.title!r})"
                )
            for t in ms.tasks:
                if t.status not in TASK_STATUS_ENUM:
                    raise RuntimeError(
                        f"Task status {t.status!r} violates CHECK (title={t.title!r})"
                    )

    return ParsedTree(workstreams=workstreams, warnings=warnings)


# ---- DB transaction ----


async def run_import_tx(
    conn: asyncpg.Connection,
    tree: ParsedTree,
    project_uuid: uuid.UUID,
    imported_by: str,
    update_project_desc: bool,
) -> Dict[str, int]:
    counts = {"workstreams": 0, "milestones": 0, "tasks": 0}
    async with conn.transaction():
        if update_project_desc:
            await conn.execute(
                "UPDATE bedrock.project SET description = $2 "
                "WHERE id = $1 AND deleted_at IS NULL",
                project_uuid, UPDATED_PROJECT_DESCRIPTION,
            )

        await conn.execute(
            "UPDATE bedrock.project_task SET deleted_at = now(), deleted_by = $2 "
            "WHERE milestone_id IN ("
            "  SELECT m.id FROM bedrock.milestone m "
            "  JOIN bedrock.workstream w ON w.id = m.workstream_id "
            "  WHERE w.project_id = $1"
            ") AND deleted_at IS NULL",
            project_uuid, imported_by,
        )
        await conn.execute(
            "UPDATE bedrock.milestone SET deleted_at = now(), deleted_by = $2 "
            "WHERE workstream_id IN (SELECT id FROM bedrock.workstream WHERE project_id = $1) "
            "AND deleted_at IS NULL",
            project_uuid, imported_by,
        )
        await conn.execute(
            "UPDATE bedrock.workstream SET deleted_at = now(), deleted_by = $2 "
            "WHERE project_id = $1 AND deleted_at IS NULL",
            project_uuid, imported_by,
        )

        for ws_row in tree.workstreams:
            ws_id = await conn.fetchval(
                "INSERT INTO bedrock.workstream (project_id, name, description, sort_order) "
                "VALUES ($1, $2, $3, $4) RETURNING id",
                project_uuid, ws_row.name, ws_row.description, ws_row.sort_order,
            )
            counts["workstreams"] += 1
            for ms in ws_row.milestones:
                ms_id = await conn.fetchval(
                    "INSERT INTO bedrock.milestone "
                    "(workstream_id, title, status, priority, owner, owner_ids, description, source_links, sort_order) "
                    "VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9) RETURNING id",
                    ws_id, ms.title, ms.status, ms.priority, ms.owner, ms.owner_ids,
                    ms.description, ms.source_links, ms.sort_order,
                )
                counts["milestones"] += 1
                for t in ms.tasks:
                    await conn.execute(
                        "INSERT INTO bedrock.project_task "
                        "(milestone_id, title, status, owner, owner_ids, deadline, start_date, "
                        " description, updates, links, sort_order) "
                        "VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11)",
                        ms_id, t.title, t.status, t.owner, t.owner_ids, t.deadline, t.start_date,
                        t.description, t.updates, t.links, t.sort_order,
                    )
                    counts["tasks"] += 1
    return counts


async def resolve_owner_ids(conn: asyncpg.Connection, tree: ParsedTree) -> Dict[str, int]:
    """Read the active-user roster, parse each milestone/task owner string,
    populate owner_ids, and reduce owner TEXT to only unmatched tokens.

    Modifies the tree in place. Returns counts of matches / residuals for the
    dry-run summary.
    """
    user_rows = await conn.fetch(
        "SELECT id, display_name FROM public.org_users "
        "WHERE COALESCE(is_active, true) = true AND display_name IS NOT NULL"
    )
    user_map = build_user_map([dict(r) for r in user_rows])

    stats = {"ms_mapped": 0, "ms_with_other": 0, "t_mapped": 0, "t_with_other": 0}
    for ws_row in tree.workstreams:
        for ms in ws_row.milestones:
            ids, other = parse_owner_string(ms.owner, user_map)
            ms.owner_ids = ids
            ms.owner = other
            if ids:
                stats["ms_mapped"] += 1
            if other:
                stats["ms_with_other"] += 1
            for t in ms.tasks:
                t_ids, t_other = parse_owner_string(t.owner, user_map)
                t.owner_ids = t_ids
                t.owner = t_other
                if t_ids:
                    stats["t_mapped"] += 1
                if t_other:
                    stats["t_with_other"] += 1
    return stats


# ---- CLI ----


def _print_summary(tree: ParsedTree, project_id: str, dry_run: bool) -> None:
    total_ws = len(tree.workstreams)
    total_ms = sum(len(w.milestones) for w in tree.workstreams)
    total_t = sum(len(m.tasks) for w in tree.workstreams for m in w.milestones)

    banner = "DRY RUN" if dry_run else "LIVE RUN (replace mode)"
    print(f"\n[{banner}] project_id={project_id}")
    print(f"  workstreams={total_ws}  milestones={total_ms}  tasks={total_t}\n")

    for ws_row in tree.workstreams:
        print(
            f"▸ [{ws_row.sort_order}] {ws_row.name!r}  "
            f"desc={len(ws_row.description)} chars  milestones={len(ws_row.milestones)}"
        )
        for ms in ws_row.milestones:
            print(
                f"    · {ms.title!r}  status={ms.status}  priority={ms.priority}  "
                f"owner={ms.owner or '-'}  src_links={len(ms.source_links)}  "
                f"desc={len(ms.description)} chars  tasks={len(ms.tasks)}"
            )
            for t in ms.tasks:
                dl = t.deadline.isoformat() if t.deadline else "-"
                sd = t.start_date.isoformat() if t.start_date else "-"
                print(
                    f"        - {t.title!r}  status={t.status}  owner={t.owner or '-'}  "
                    f"deadline={dl}  start={sd}  links={len(t.links)}  updates={len(t.updates)} chars"
                )

    if tree.warnings:
        print("\nWarnings:")
        for w in tree.warnings:
            print(f"  ! {w}")
    print()


def _confirm(prompt: str) -> bool:
    try:
        return input(f"{prompt} [y/N]: ").strip().lower() in ("y", "yes")
    except EOFError:
        return False


async def _amain(args: argparse.Namespace) -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    path = Path(args.input).expanduser().resolve()
    if not path.exists():
        print(f"ERROR: input file not found: {path}", file=sys.stderr)
        return 2

    tree = parse_workbook(path)

    # Owner-ID resolution needs the DB too — even in dry-run, we connect
    # read-only so the summary can show the real UUID mapping + residual Other.
    dsn = (os.getenv("DATABASE_URL") or "").strip()
    if not dsn:
        print(
            "ERROR: DATABASE_URL is not set. Refusing to run blind — set it to "
            "the target DB (staging) and retry.",
            file=sys.stderr,
        )
        return 3

    parsed_dsn = urlparse(dsn)
    masked = (
        f"{parsed_dsn.scheme}://{parsed_dsn.username}:****@"
        f"{parsed_dsn.hostname}:{parsed_dsn.port}{parsed_dsn.path}"
    )
    print(f"Target DB: {masked}")

    try:
        project_uuid = uuid.UUID(args.project_id)
    except ValueError:
        print(f"ERROR: --project-id {args.project_id!r} is not a valid UUID", file=sys.stderr)
        return 4

    conn = await asyncpg.connect(dsn)
    try:
        owner_stats = await resolve_owner_ids(conn, tree)
        _print_summary(tree, args.project_id, dry_run=args.dry_run)
        print(
            "Owner resolution:  "
            f"milestones_mapped={owner_stats['ms_mapped']}  "
            f"milestones_with_other={owner_stats['ms_with_other']}  "
            f"tasks_mapped={owner_stats['t_mapped']}  "
            f"tasks_with_other={owner_stats['t_with_other']}"
        )

        if args.dry_run:
            print("\nDry run only. No DB writes performed.")
            return 0

        proj = await conn.fetchrow(
            "SELECT id, name, description, owner_email "
            "FROM bedrock.project WHERE id = $1 AND deleted_at IS NULL",
            project_uuid,
        )
        if not proj:
            print(
                f"ERROR: project {args.project_id} not found or soft-deleted "
                "in bedrock.project on the target DB.",
                file=sys.stderr,
            )
            return 5

        existing = await conn.fetchrow(
            """
            SELECT
                (SELECT count(*) FROM bedrock.workstream
                    WHERE project_id = $1 AND deleted_at IS NULL) AS ws_count,
                (SELECT count(*) FROM bedrock.milestone m
                    JOIN bedrock.workstream w ON w.id = m.workstream_id
                    WHERE w.project_id = $1 AND m.deleted_at IS NULL) AS ms_count,
                (SELECT count(*) FROM bedrock.project_task t
                    JOIN bedrock.milestone m ON m.id = t.milestone_id
                    JOIN bedrock.workstream w ON w.id = m.workstream_id
                    WHERE w.project_id = $1 AND t.deleted_at IS NULL) AS task_count
            """,
            project_uuid,
        )
        print(
            f"Target project: {proj['name']!r} (owner={proj['owner_email']!r})\n"
            f"  existing tree: {existing['ws_count']} workstreams, "
            f"{existing['ms_count']} milestones, {existing['task_count']} tasks"
        )
        new_ms = sum(len(w.milestones) for w in tree.workstreams)
        new_t = sum(len(m.tasks) for w in tree.workstreams for m in w.milestones)
        print(
            f"  will replace with: {len(tree.workstreams)} workstreams, "
            f"{new_ms} milestones, {new_t} tasks"
        )
        if args.update_project_desc:
            print(f"  project.description will be updated to:\n    {UPDATED_PROJECT_DESCRIPTION!r}")

        if not args.yes and not _confirm(
            f"\nSoft-delete existing AIJI tree and insert the new one on {masked}?"
        ):
            print("Aborted.")
            return 0

        counts = await run_import_tx(
            conn=conn,
            tree=tree,
            project_uuid=project_uuid,
            imported_by=args.imported_by,
            update_project_desc=args.update_project_desc,
        )
        print(f"\nImport complete. Inserted: {counts}")

        post = await conn.fetchrow(
            """
            SELECT
                (SELECT count(*) FROM bedrock.workstream
                    WHERE project_id = $1 AND deleted_at IS NULL) AS ws_count,
                (SELECT count(*) FROM bedrock.milestone m
                    JOIN bedrock.workstream w ON w.id = m.workstream_id
                    WHERE w.project_id = $1 AND m.deleted_at IS NULL) AS ms_count,
                (SELECT count(*) FROM bedrock.project_task t
                    JOIN bedrock.milestone m ON m.id = t.milestone_id
                    JOIN bedrock.workstream w ON w.id = m.workstream_id
                    WHERE w.project_id = $1 AND t.deleted_at IS NULL) AS task_count
            """,
            project_uuid,
        )
        print(
            f"Post-import live rows: {post['ws_count']} workstreams, "
            f"{post['ms_count']} milestones, {post['task_count']} tasks"
        )
        return 0
    finally:
        await conn.close()


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Import AIJI Project Tracker v12 into bedrock.project tree."
    )
    parser.add_argument("--input", default=DEFAULT_INPUT_PATH, help="Path to the .xlsx file")
    parser.add_argument("--project-id", default=DEFAULT_PROJECT_ID, help="bedrock.project UUID")
    parser.add_argument(
        "--imported-by", default=DEFAULT_IMPORTED_BY,
        help="Email recorded in deleted_by for the soft-delete audit trail",
    )
    parser.add_argument("--dry-run", action="store_true", help="Parse + summarize, no DB writes")
    parser.add_argument("--yes", "-y", action="store_true", help="Skip the confirmation prompt")
    parser.add_argument(
        "--update-project-desc", action="store_true",
        help="Also update bedrock.project.description to the Excel-derived summary",
    )
    args = parser.parse_args(argv)
    return asyncio.run(_amain(args))


if __name__ == "__main__":
    sys.exit(main())
